"""
extract_financial_statements.py
从年报 PDF 中提取合并资产负债表 / 合并利润表 / 合并现金流量表，写入 Excel。

设计要点：
- 复用 PDFChapterExtractor 解析 TOC 与表格，但避开其有缺陷的跨页合并逻辑
- 每次按页打开独立 fitz 文档，避免 PyMuPDF find_tables 在多次调用时返回不同结果
- 通过表头文本（"1、合并资产负债表"等）精确识别每张报表的起点
- 通过"下一报表所在页"确定报表结束边界，避免把母公司报表混入
"""
from __future__ import annotations

import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 6 张主要报表在 TOC 中的标题（按出现顺序）
MAIN_STATEMENT_TITLES: List[str] = [
    "合并资产负债表",
    "母公司资产负债表",
    "合并利润表",
    "母公司利润表",
    "合并现金流量表",
    "母公司现金流量表",
]

# 我们只需要导出合并报表的 3 张
TARGET_STATEMENTS: List[str] = [
    "合并资产负债表",
    "合并利润表",
    "合并现金流量表",
]

# 表头识别关键字（含前缀编号 + 报表名）
HEADER_KEYWORDS: Dict[str, List[str]] = {
    "合并资产负债表": ["合并资产负债表"],
    "母公司资产负债表": ["母公司资产负债表"],
    "合并利润表": ["合并利润表"],
    "母公司利润表": ["母公司利润表"],
    "合并现金流量表": ["合并现金流量表"],
    "母公司现金流量表": ["母公司现金流量表"],
}


def _is_header_match(header_text: str, statement_title: str) -> bool:
    """判断 header_text 是否包含指定报表的标识。"""
    if not header_text:
        return False
    keywords = HEADER_KEYWORDS.get(statement_title, [statement_title])
    return any(kw in header_text for kw in keywords)


def find_section_ranges(pdf_path: str) -> Dict[str, Tuple[int, int]]:
    """从 PDF 的 TOC 中找出 6 张主要报表所在页码范围（页码 0-based）。

    优先使用 fitz 内置 TOC，否则使用 toc_fallback_parser 解析。
    """
    from src.tools.toc_fallback_parser import parse_toc_fallback

    doc = fitz.open(pdf_path)
    toc = list(doc.get_toc())
    if not toc:
        toc = parse_toc_fallback(pdf_path)
    page_heights = [page.rect.height for page in doc]
    doc.close()

    # 复用 find_section_range 的等价逻辑
    ranges: Dict[str, Tuple[int, int]] = {}

    def _find(title: str) -> Optional[Tuple[int, int]]:
        # 优先按 level=3 精确匹配
        for i, entry in enumerate(toc):
            level, entry_title, page_num = entry
            if level == 3 and title in entry_title:
                start = page_num - 1
                # 找下一个 level<=3 的 entry 作为结束
                end = len(page_heights) - 1
                for j in range(i + 1, len(toc)):
                    if toc[j][0] <= 3:
                        end = toc[j][2] - 1
                        break
                return start, end
        # fallback: 模糊匹配
        for i, entry in enumerate(toc):
            level, entry_title, page_num = entry
            if title in entry_title:
                start = page_num - 1
                end = len(page_heights) - 1
                for j in range(i + 1, len(toc)):
                    if toc[j][0] <= entry[0]:
                        end = toc[j][2] - 1
                        break
                return start, end
        return None

    for title in MAIN_STATEMENT_TITLES:
        rng = _find(title)
        if rng is not None:
            ranges[title] = rng
    return ranges


def _get_page_top_text(doc: "fitz.Document", page_num: int, table_bbox: Tuple[float, float, float, float]) -> str:
    """获取页面上指定表格上方的若干行文本（作为表头识别来源）。"""
    page = doc[page_num]
    page_dict = page.get_text("dict")
    table_top = table_bbox[1]

    lines: List[Tuple[float, str]] = []
    for block in page_dict["blocks"]:
        if block["type"] != 0:
            continue
        for line in block["lines"]:
            line_text = "".join(
                span["text"].strip() for span in line["spans"] if span["text"].strip()
            )
            if not line_text:
                continue
            y_top = line["spans"][0]["bbox"][1] if line["spans"] else 0
            if y_top < table_top:
                lines.append((y_top, line_text))
    lines.sort(key=lambda x: x[0])
    # 取最接近表格的 5 行
    return "\n".join(text for _, text in lines[-5:])


def _extract_tables_on_page(pdf_path: str, page_num: int) -> List[Dict]:
    """从指定 PDF 页面提取所有表格。

    为避免 PyMuPDF find_tables 的内部状态问题，每次按页打开独立文档。
    """
    doc = fitz.open(pdf_path)
    try:
        page = doc[page_num]
        page_height = page.rect.height
        tabs = list(page.find_tables(strategy="lines_strict"))
        tables: List[Dict] = []
        for tab in tabs:
            if not tab.header or not tab.cells:
                continue
            header_text = _get_page_top_text(doc, page_num, tab.bbox)
            try:
                data = tab.extract()
            except Exception:
                data = []
            tables.append(
                {
                    "page": page_num,
                    "bbox": tab.bbox,
                    "header_text": header_text,
                    "data": data,
                }
            )
        # 按 y_top 排序
        tables.sort(key=lambda t: t["bbox"][1])
        return tables
    finally:
        doc.close()


def extract_consolidated_statement(
    pdf_path: str,
    statement_title: str,
    section_ranges: Dict[str, Tuple[int, int]],
) -> List[List[str]]:
    """提取指定合并报表的所有数据行。

    算法：
    1. 根据 TOC 拿到该报表的起始页
    2. 结束页 = 下一张报表（母公司报表）起始页 - 1
       （因为母公司报表往往在该合并报表所在页的下一页开始）
    3. 从起始页到结束页，提取所有表格
    4. 跳过表头标识属于其他报表（尤其是下一张报表）的表格
    5. 从剩余表格中找到"该报表的起点"（首个表头包含本报表名的表格）
    6. 拼接所有相关表格的数据行

    Returns: 二维列表（行 x 列），包含列头。
    """
    if statement_title not in section_ranges:
        raise ValueError(f"PDF 中未找到报表: {statement_title}")

    start_page, end_page = section_ranges[statement_title]

    # 包含到该 section 的实际结束页；与下一张报表共页时，由表头匹配规则排除混入的母公司表
    upper_page = end_page

    # 收集页内表格
    page_tables: Dict[int, List[Dict]] = {}
    for page_num in range(start_page, upper_page + 1):
        page_tables[page_num] = _extract_tables_on_page(pdf_path, page_num)

    # 找出"该报表的起点"表格
    # 候选起始表：表头包含本报表标题
    # 在 [start_page, upper_page] 范围内，最靠前的即为起点
    candidate_start = None
    for page_num in sorted(page_tables.keys()):
        for idx, tab in enumerate(page_tables[page_num]):
            if _is_header_match(tab["header_text"], statement_title):
                candidate_start = (page_num, idx)
                break
        if candidate_start:
            break

    # 兜底：取起始页的第一个表格
    if candidate_start is None:
        for page_num in sorted(page_tables.keys()):
            if page_tables[page_num]:
                candidate_start = (page_num, 0)
                break

    if candidate_start is None:
        raise ValueError(f"未在 PDF 中找到报表 {statement_title} 的起始表")

    # 收集"起点及其之后"的表格，跳过其他报表的表
    start_pg, start_idx = candidate_start

    rows: List[List[str]] = []
    found_start = False
    other_statement_titles = set(MAIN_STATEMENT_TITLES) - {statement_title}

    for page_num in sorted(page_tables.keys()):
        for idx, tab in enumerate(page_tables[page_num]):
            if not found_start:
                if (page_num, idx) == candidate_start:
                    found_start = True
                else:
                    continue

            # 跳过其他报表的表头（特别是下一张报表）
            if any(
                _is_header_match(tab["header_text"], other)
                for other in other_statement_titles
            ):
                continue

            if tab["data"]:
                rows.extend(tab["data"])

    if not rows:
        raise ValueError(f"报表 {statement_title} 提取为空")

    return rows


def clean_rows(rows: List[List[str]]) -> List[List[str]]:
    """清洗表格行：去除完全空行、规范化单元格。"""
    cleaned: List[List[str]] = []
    for row in rows:
        if not row:
            continue
        # 转字符串并 strip
        normalized = [
            ("" if cell is None else str(cell).replace("\n", " ").strip())
            for cell in row
        ]
        # 跳过全空行
        if all(c == "" for c in normalized):
            continue
        cleaned.append(normalized)
    return cleaned


def write_excel(
    output_path: str,
    statement_data: Dict[str, List[List[str]]],
) -> None:
    """将三张报表写入 Excel，每个报表一个 sheet。"""
    wb = Workbook()
    # 删除默认 sheet
    default_ws = wb.active
    wb.remove(default_ws)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for title in TARGET_STATEMENTS:
        rows = statement_data.get(title, [])
        ws = wb.create_sheet(title=title)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
                else:
                    # 数字列右对齐，文本列左对齐
                    if re.match(r"^[-+]?[\d,]+\.?\d*$", str(value).strip()):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = left

        # 设置列宽
        col_count = max((len(r) for r in rows), default=3)
        ws.column_dimensions["A"].width = 38
        for c in range(2, col_count + 1):
            ws.column_dimensions[get_column_letter(c)].width = 22
        ws.freeze_panes = "B2"

    wb.save(output_path)


def process_pdf(pdf_path: str, output_dir: str) -> str:
    """处理单个 PDF，返回生成的 Excel 路径。"""
    pdf_path = os.path.abspath(pdf_path)
    pdf_name = Path(pdf_path).stem  # 例如 东阿阿胶_000423_2024
    output_path = os.path.join(output_dir, f"{pdf_name}.xlsx")

    print(f"\n=== 处理 {pdf_name} ===")
    section_ranges = find_section_ranges(pdf_path)
    print(f"找到的报表页码范围（0-based）:")
    for title, rng in section_ranges.items():
        print(f"  {title}: pages {rng[0]+1}-{rng[1]+1}")

    statement_data: Dict[str, List[List[str]]] = {}
    for title in TARGET_STATEMENTS:
        try:
            rows = extract_consolidated_statement(pdf_path, title, section_ranges)
            cleaned = clean_rows(rows)
            statement_data[title] = cleaned
            print(f"  ✔ {title}: {len(cleaned)} 行")
        except Exception as e:
            print(f"  ✘ {title} 提取失败: {e}")
            statement_data[title] = []

    write_excel(output_path, statement_data)
    print(f"已生成: {output_path}")
    return output_path


def main():
    if len(sys.argv) < 2:
        print("用法: python extract_financial_statements.py <pdf_or_dir> [output_dir]")
        sys.exit(1)

    input_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "data/000423_label"
    os.makedirs(output_dir, exist_ok=True)

    pdfs: List[str] = []
    if os.path.isdir(input_path):
        pdfs = sorted(
            os.path.join(input_path, f)
            for f in os.listdir(input_path)
            if f.lower().endswith(".pdf")
        )
    else:
        pdfs = [input_path]

    results: List[str] = []
    for pdf in pdfs:
        try:
            results.append(process_pdf(pdf, output_dir))
        except Exception as e:
            print(f"处理 {pdf} 时出错: {e}")

    print(f"\n共处理 {len(results)} 个文件，输出目录: {output_dir}")


if __name__ == "__main__":
    main()